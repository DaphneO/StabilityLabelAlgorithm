from pathlib import Path
from typing import Union

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from ..argumentation_theory.literal import Literal
from ..argumentation_theory.queryable import Queryable
from ..argumentation_theory.rule import Rule


class ArgumentationSystemXLSXReader:
    def __init__(self, path_to_xls: Union[Path, str]):
        wb_sheet_names = load_workbook(path_to_xls, read_only=True).sheetnames

        self._literal_df = pd.read_excel(path_to_xls, sheet_name='Literals')
        self._rule_df = pd.read_excel(path_to_xls, sheet_name='Rules')
        self._query_df = pd.read_excel(path_to_xls, sheet_name='Queries', index_col='Query')

        if 'Contraries' in wb_sheet_names:
            self._contrary_df = pd.read_excel(path_to_xls, sheet_name='Contraries')
        else:
            self._contrary_df = None
        if 'RulePreferences' in wb_sheet_names:
            _rule_preference_df = pd.read_excel(path_to_xls, sheet_name='RulePreferences')
        else:
            _rule_preference_df = None
        if 'Positions' in wb_sheet_names:
            self._position_df = pd.read_excel(path_to_xls, sheet_name='Positions')
        else:
            self._position_df = None
        if 'About' in wb_sheet_names:
            self.about_text = pd.read_excel(path_to_xls, sheet_name='About', index_col=False, header=None).loc[0, 0]
        else:
            self.about_text = ''

        try:
            self._get_language()
            self._get_rules()
            self._connect_parents_and_children()
            self.rule_preference_matrix = None
            self._add_rule_preferences(_rule_preference_df)
        except Exception as exception:
            raise ImportError('Could not load Argumentation System.' + str(type(exception)))

        self.source_path = path_to_xls

    def _get_language(self):
        """
        Get the logical language from the argumentation system. A literal l is in the
        logical language if and only if its negation is in the logical language as well.

        :return: Dict (indexed on name string) of literals in the logical language
        """
        self.language = dict()
        self.topic_literals = []

        for index, row in self._literal_df.iterrows():
            if row['Observable'] == 'y':
                query = self._query_df.loc[row['Literal'], 'NL(Query)']
                query_explanation = self._query_df.loc[row['Literal'], 'EXP(Query)']
                if 'Priority' in self._query_df.keys():
                    priority = self._query_df.loc[row['Literal'], 'Priority']
                else:
                    priority = 0
                literal = Queryable(row['Literal'], row['NL(True)'], row['NL(Unknown)'],
                                    query, query_explanation, priority)
                literal_neg = Queryable('~' + row['Literal'], row['NL(False)'], row['NL(Unknown)'],
                                        query, query_explanation, priority)
            else:
                literal = Literal(row['Literal'], row['NL(True)'], row['NL(Unknown)'], False)
                literal_neg = Literal('~' + row['Literal'], row['NL(False)'], row['NL(Unknown)'], False)

            if 'Topic' in row and row['Topic'] == 'y':
                self.topic_literals.append(literal)

            literal.contraries.append(literal_neg)
            literal_neg.contraries.append(literal)
            literal.negation = literal_neg
            literal_neg.negation = literal
            self.language[str(literal)] = literal
            self.language[str(literal_neg)] = literal_neg

        # Add additional contraries
        if self._contrary_df is not None:
            for index, row in self._contrary_df.iterrows():
                literal_str = row['Literal']
                contraries_str = row['Contraries'].split(',')
                for contrary_str in contraries_str:
                    self.language[literal_str].contraries.append(self.language[contrary_str])

        # Add positions if given
        if self._position_df is not None:
            for index, row in self._position_df.iterrows():
                literal_str = row['Literal']
                x_pos = row['X']
                y_pos = row['Y']
                self.language[literal_str].position = (x_pos, y_pos)

    def _get_rules(self):
        """
        Get all rules (R) from the argumentation system.

        :return: Rule list
        """
        self.rules = []
        for index, row in self._rule_df.iterrows():
            ants_str = [ant.strip() for ant in row['Antecedents'].split(',')]
            cons_str = row['Consequent'].strip()
            if any([ant not in self.language for ant in ants_str]) or cons_str not in self.language:
                raise KeyError(
                    'Not each literal in ' + str(ants_str) + ' or ' + cons_str + ' was in logical language.')
            ants = {self.language[literal_str] for literal_str in ants_str}
            cons = self.language[cons_str]
            exp = row['NL(Rule)']

            if 'ID' in self._rule_df.keys():
                rule_id = row['ID']
            else:
                rule_id = index

            self.rules.append(Rule(rule_id, ants, cons, exp))

    def _connect_parents_and_children(self):
        for rule in self.rules:
            rule.consequent.children.append(rule)
            for child in rule.antecedents:
                child.parents.append(rule)

    @staticmethod
    def _get_updated_preference(old_value: str, new_value: str):
        if old_value not in ['?', '<', '=', '>']:
            raise ImportError('{} is not a suitable operator string. '
                              'Use "?", "<", "=" or ">" instead.'.format(old_value))
        if new_value not in ['?', '<', '=', '>']:
            raise ImportError('{} is not a suitable operator string. '
                              'Use "?", "<", "=" or ">" instead.'.format(new_value))

        if old_value == '?':
            return new_value
        if new_value == '?':
            return old_value

        # From here: both old and new value are in ['<', '=', '>']
        if old_value == '<':
            if new_value in ['=', '>']:
                raise ImportError('Rule ordering is not correct.')
            return old_value
        if old_value == '=':
            if new_value in ['<', '>']:
                raise ImportError('Rule ordering is not correct.')
            return old_value
        if new_value in ['<', '=']:
            raise ImportError('Rule ordering is not correct.')
        return old_value

    @staticmethod
    def _try_transitivity(operator_a: str, operator_b: str) -> str:
        if operator_a == '>' and operator_b == '>':
            return '>'
        if operator_a == '<' and operator_b == '<':
            return '<'
        if operator_a == '=':
            return operator_b
        if operator_b == '=':
            return operator_a
        return ''

    @staticmethod
    def _inverse_operator(operator: str) -> str:
        return {'?': '?', '<': '>', '=': '=', '>': '<'}[operator]

    def _set_preference(self, rule_index_a: int, preference: str, rule_index_b):
        if rule_index_a < rule_index_b:
            self.rule_preference_matrix.loc[rule_index_a, rule_index_b] = preference
        else:
            self.rule_preference_matrix.loc[rule_index_b, rule_index_a] = self._inverse_operator(preference)

    def _get_preference(self, rule_index_a: int, rule_index_b: int) -> str:
        if rule_index_a < rule_index_b:
            return self.rule_preference_matrix.loc[rule_index_a, rule_index_b]
        else:
            return self._inverse_operator(self.rule_preference_matrix.loc[rule_index_b, rule_index_a])

    def _add_rule_preferences(self, rule_preference_df) -> None:
        # Initialise preferences: each rule is as strong as itself, for other rules we do not know yet.
        array = np.full((len(self.rules), len(self.rules)), '?')
        np.fill_diagonal(array, '=')
        self.rule_preference_matrix = pd.DataFrame(array, index=[r.id for r in self.rules],
                                                   columns=[r.id for r in self.rules])

        # Update self.rule_preference_matrix with information from rule_preference_df (if it exists)
        if rule_preference_df is not None:
            changed = []
            for index, row in rule_preference_df.iterrows():
                try:
                    rule_a = int(row['RuleID1'])
                    rule_b = int(row['RuleID2'])
                    operator: str = row['Operator']
                except ValueError:
                    raise ImportError('Rule preference tab should code rules with indices and operators with a str.')

                old_a_b = self._get_preference(rule_a, rule_b)
                new_a_b = self._get_updated_preference(old_a_b, operator)

                if old_a_b != new_a_b:
                    self._set_preference(rule_a, new_a_b, rule_b)
                    changed.append((rule_a, rule_b))

            # Iteratively apply transitivity
            while changed:
                changed_new = []

                for rule_a, rule_b in changed:
                    for rule in self.rules:
                        if rule.id not in [rule_a, rule_b]:
                            # Obtain the old preferences between the new rule and rule a and b
                            old_r_a = self._get_preference(rule.id, rule_a)
                            old_r_b = self._get_preference(rule.id, rule_b)

                            new_r_a = self._try_transitivity(old_r_b, self._get_preference(rule_b, rule_a))
                            if new_r_a != '' and old_r_a != new_r_a:
                                self._set_preference(rule.id, new_r_a, rule_a)
                                old_r_a = new_r_a
                                changed_new.append((rule.id, rule_a))

                            new_r_b = self._try_transitivity(old_r_a, self._get_preference(rule_a, rule_b))
                            if new_r_b != '' and old_r_b != new_r_b:
                                self._set_preference(rule.id, new_r_b, rule_b)
                                changed_new.append((rule.id, rule_b))

                changed = changed_new
